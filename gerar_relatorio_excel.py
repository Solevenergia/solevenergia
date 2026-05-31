"""
Gera relatório de UCs em formato Excel.
"""

import json
import sys
from pathlib import Path

# Fix encoding
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
except ImportError:
    print("Erro: openpyxl nao instalado. Execute: pip install openpyxl")
    sys.exit(1)

try:
    from db import tb_mapa_uc_para_usina
except ImportError:
    print("Erro: nao conseguiu importar db.py")
    sys.exit(1)


def carregar_ucs_equatorial(arquivo_json: str = "ucs_equatorial.json") -> list:
    """Carrega UCs do arquivo JSON"""
    try:
        with open(arquivo_json, "r", encoding="utf-8") as f:
            dados = json.load(f)
        return [uc["valor"].strip() for uc in dados if uc.get("valor")]
    except Exception as e:
        print(f"Erro ao carregar {arquivo_json}: {e}")
        return []


def carregar_ucs_solev() -> list:
    """Carrega UCs do SOLEV"""
    try:
        mapa = tb_mapa_uc_para_usina()
        return list(mapa.keys()) if mapa else []
    except Exception as e:
        print(f"Erro ao carregar UCs do SOLEV: {e}")
        return []


def aplicar_estilo_header(ws, row_num):
    """Aplica estilo de header a uma linha"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[row_num]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border


def aplicar_estilo_dados(ws, inicio_row, fim_row):
    """Aplica estilo aos dados"""
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=inicio_row, max_row=fim_row):
        for cell in row:
            cell.border = border
            if cell.column == 1:  # Coluna de numero
                cell.alignment = Alignment(horizontal="center")
            cell.alignment = Alignment(vertical="center")


def main():
    print("Carregando dados...")

    # Carrega dados
    ucs_equatorial = carregar_ucs_equatorial()
    ucs_solev = carregar_ucs_solev()

    if not ucs_equatorial:
        print("Erro: nenhuma UC foi carregada")
        sys.exit(1)

    set_equatorial = set(ucs_equatorial)
    set_solev = set(ucs_solev)

    faltando = sorted(set_equatorial - set_solev)
    extras = sorted(set_solev - set_equatorial)
    ja_cadastradas = sorted(set_equatorial & set_solev)

    # Cria workbook
    print("Criando arquivo Excel...")
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    # Sheet 1: Resumo
    ws_resumo["A1"] = "RELATORIO DE CADASTRO DE UCs - EQUATORIAL GOIAS"
    ws_resumo["A1"].font = Font(bold=True, size=14)
    ws_resumo.merge_cells("A1:B1")

    ws_resumo["A3"] = "Data Extracao:"
    ws_resumo["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws_resumo["A4"] = "CPF Titular:"
    ws_resumo["B4"] = "01873853190"

    ws_resumo["A6"] = "RESUMO GERAL"
    ws_resumo["A6"].font = Font(bold=True, size=12)

    resumo_data = [
        ("Total de UCs no Equatorial", len(set_equatorial)),
        ("Total de UCs cadastradas em SOLEV", len(set_solev)),
        ("Ja cadastradas (match)", len(ja_cadastradas)),
        ("FALTANDO CADASTRAR", len(faltando)),
        ("EXTRAS (em SOLEV mas nao em Equatorial)", len(extras)),
    ]

    row = 7
    for desc, valor in resumo_data:
        ws_resumo[f"A{row}"] = desc
        ws_resumo[f"B{row}"] = valor
        ws_resumo[f"B{row}"].alignment = Alignment(horizontal="center")
        if "FALTANDO" in desc or "EXTRAS" in desc:
            ws_resumo[f"A{row}"].font = Font(bold=True, color="C00000")
            ws_resumo[f"B{row}"].font = Font(bold=True, color="C00000")
        row += 1

    ws_resumo.column_dimensions["A"].width = 40
    ws_resumo.column_dimensions["B"].width = 20

    # Sheet 2: UCs Faltando
    ws_faltando = wb.create_sheet("UCs Faltando (93)")
    ws_faltando["A1"] = "#"
    ws_faltando["B1"] = "UC (15 digitos)"
    aplicar_estilo_header(ws_faltando, "1:1")

    for i, uc in enumerate(faltando, 1):
        ws_faltando[f"A{i+1}"] = i
        ws_faltando[f"B{i+1}"] = uc

    aplicar_estilo_dados(ws_faltando, 2, len(faltando) + 1)
    ws_faltando.column_dimensions["A"].width = 8
    ws_faltando.column_dimensions["B"].width = 20

    # Sheet 3: UCs Ja Cadastradas
    ws_cadastradas = wb.create_sheet("Ja Cadastradas (85)")
    ws_cadastradas["A1"] = "#"
    ws_cadastradas["B1"] = "UC (15 digitos)"
    aplicar_estilo_header(ws_cadastradas, "1:1")

    for i, uc in enumerate(ja_cadastradas, 1):
        ws_cadastradas[f"A{i+1}"] = i
        ws_cadastradas[f"B{i+1}"] = uc

    aplicar_estilo_dados(ws_cadastradas, 2, len(ja_cadastradas) + 1)
    ws_cadastradas.column_dimensions["A"].width = 8
    ws_cadastradas.column_dimensions["B"].width = 20

    # Sheet 4: UCs Extras
    ws_extras = wb.create_sheet("Extras - Revisar (9)")
    ws_extras["A1"] = "#"
    ws_extras["B1"] = "UC (15 digitos)"
    aplicar_estilo_header(ws_extras, "1:1")

    for i, uc in enumerate(extras, 1):
        ws_extras[f"A{i+1}"] = i
        ws_extras[f"B{i+1}"] = uc

    aplicar_estilo_dados(ws_extras, 2, len(extras) + 1)
    ws_extras.column_dimensions["A"].width = 8
    ws_extras.column_dimensions["B"].width = 20

    # Sheet 5: Todas as UCs Equatorial
    ws_todas = wb.create_sheet("Todas Equatorial (178)")
    ws_todas["A1"] = "#"
    ws_todas["B1"] = "UC (15 digitos)"
    ws_todas["C1"] = "Status"
    aplicar_estilo_header(ws_todas, "1:1")

    for i, uc in enumerate(ucs_equatorial, 1):
        ws_todas[f"A{i+1}"] = i
        ws_todas[f"B{i+1}"] = uc

        if uc in ja_cadastradas:
            ws_todas[f"C{i+1}"] = "Cadastrada"
            ws_todas[f"C{i+1}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws_todas[f"C{i+1}"] = "Faltando"
            ws_todas[f"C{i+1}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    aplicar_estilo_dados(ws_todas, 2, len(ucs_equatorial) + 1)
    ws_todas.column_dimensions["A"].width = 8
    ws_todas.column_dimensions["B"].width = 20
    ws_todas.column_dimensions["C"].width = 15

    # Salva arquivo
    output_file = "Relatorio_UCs_Equatorial.xlsx"
    wb.save(output_file)
    print(f"\nOK: Relatorio gerado com sucesso!")
    print(f"    Arquivo: {output_file}")
    print(f"    Tamanho: {Path(output_file).stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
