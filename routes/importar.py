"""
Blueprint: Importar
Rotas: /importar, /importar/modelo
Helpers: _tipo_fornecimento, _importar_formato_padrao, _importar_formato_controle
"""
import os
import traceback
from datetime import datetime as _dt
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file

bp = Blueprint('importar', __name__)

_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_FOLDER = os.path.join(_PROJECT_DIR, "uploads")


def _tipo_fornecimento(val):
    """Normaliza tipo de fornecimento para o padrao tb_clientes."""
    if not val:
        return "MONOFASICO"
    v = str(val).upper().strip()
    if v in ("MONO", "MONOFASICO"):
        return "MONOFASICO"
    if v in ("BI", "BIFASICO"):
        return "BIFASICO"
    if v in ("TRI", "TRIFASICO", "TR"):
        return "TRIFASICO"
    return "MONOFASICO"


def _importar_formato_padrao(xlsx_path):
    """Importa planilha no formato padrao (clientes em linhas) -> salva em tb_clientes.

    Colunas esperadas (a partir da linha 4, linha 3 = exemplo):
     1  UC                 cod_uc              (obrigatorio)
     2  NOME               desc_nome           (obrigatorio)
     3  CPF                desc_cpf
     4  TELEFONE           desc_telefone
     5  EMAIL              desc_email
     6  APELIDO            desc_apelido
     7  LOGRADOURO         desc_logradouro
     8  NUMERO             desc_numero
     9  COMPLEMENTO        desc_complemento
    10  BAIRRO             desc_setor
    11  CEP                cod_cep
    12  CIDADE             desc_cidade
    13  ESTADO             desc_estado
    14  TIPO FORNECIMENTO  tp_fornecimento
    15  TARIFA (R$/kWh)    vlr_tarifa_sem
    16  DESCONTO (%)       pct_desconto
    17  ID USINA           id_usina (vinculo cliente-usina)
    """
    import openpyxl
    from db import tb_save_cliente, tb_save_endereco, tb_get_cliente_por_uc, tb_save_cliente_usina
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    resultados = []

    def _safe_float(val, default=0.0):
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, _dt):
            return default
        s = str(val).strip().replace(",", ".").replace("R$", "").replace(" ", "")
        try:
            return float(s)
        except Exception:
            return default

    def _safe_str(val):
        if val is None:
            return ""
        if isinstance(val, _dt):
            return val.strftime("%d/%m/%Y")
        return str(val).strip()

    def _safe_int(val):
        if val is None:
            return None
        try:
            return int(str(val).strip())
        except Exception:
            return None

    # Dados comecam na linha 4 (1=cabecalho, 2=descricoes, 3=exemplo)
    for row_idx in range(4, ws.max_row + 1):
        uc = _safe_str(ws.cell(row=row_idx, column=1).value).replace(" ", "").replace(".", "")
        if not uc or uc == "None":
            continue
        nome = _safe_str(ws.cell(row=row_idx, column=2).value).upper()
        if not nome:
            continue

        tarifa = _safe_float(ws.cell(row=row_idx, column=15).value, 1.125214)
        desc   = _safe_float(ws.cell(row=row_idx, column=16).value, 20)
        if desc > 1:
            desc = desc / 100
        id_usina = _safe_int(ws.cell(row=row_idx, column=17).value)

        ja_existe = tb_get_cliente_por_uc(uc) is not None

        cli_salvo = tb_save_cliente({
            "cod_uc":          uc,
            "desc_nome":       nome,
            "desc_cpf":        _safe_str(ws.cell(row=row_idx, column=3).value),
            "desc_telefone":   _safe_str(ws.cell(row=row_idx, column=4).value),
            "desc_email":      _safe_str(ws.cell(row=row_idx, column=5).value),
            "desc_apelido":    _safe_str(ws.cell(row=row_idx, column=6).value),
            "tp_fornecimento": _tipo_fornecimento(ws.cell(row=row_idx, column=14).value),
            "vlr_tarifa_sem":  tarifa,
            "pct_desconto":    desc,
        })
        id_cliente = cli_salvo.get("id_cliente") if cli_salvo else None
        if id_cliente:
            end = {
                "desc_logradouro": _safe_str(ws.cell(row=row_idx, column=7).value).upper(),
                "desc_numero":     _safe_str(ws.cell(row=row_idx, column=8).value),
                "desc_complemento":_safe_str(ws.cell(row=row_idx, column=9).value),
                "desc_setor":      _safe_str(ws.cell(row=row_idx, column=10).value).upper(),
                "cod_cep":         _safe_str(ws.cell(row=row_idx, column=11).value),
                "desc_cidade":     _safe_str(ws.cell(row=row_idx, column=12).value),
                "desc_estado":     _safe_str(ws.cell(row=row_idx, column=13).value).upper()[:2],
            }
            if any(v for v in end.values()):
                tb_save_endereco(id_cliente, end)
            if id_usina:
                try:
                    tb_save_cliente_usina(id_cliente, id_usina)
                except Exception:
                    pass  # vinculo duplicado ou usina inexistente — nao bloqueia importacao

        resultados.append({"uc": uc, "nome": nome,
                           "status": "Atualizado" if ja_existe else "Novo", "ok": True})
    return resultados


def _importar_formato_controle(xlsx_path):
    """Importa planilha no formato CONTROLE (clientes em colunas) -> salva em tb_clientes."""
    import openpyxl
    from db import tb_save_cliente, tb_save_endereco, tb_get_cliente_por_uc
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[-1]]
    resultados = []

    def _safe_str(val):
        if val is None:
            return ""
        if isinstance(val, _dt):
            return val.strftime("%d/%m/%Y")
        return str(val).strip()

    def _safe_float(val, default=0.0):
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, _dt):
            return default
        s = str(val).strip().replace(",", ".").replace("R$", "").replace(" ", "")
        try:
            return float(s)
        except Exception:
            return default

    # Mapeia campos por linha (coluna B ou C contem os labels)
    label_map = {}
    for row_idx in range(1, min(50, ws.max_row + 1)):
        for col in [2, 3]:
            val = str(ws.cell(row=row_idx, column=col).value or "").strip().upper()
            if not val:
                continue
            if "NOME" in val:
                label_map["nome"] = row_idx
            elif "CPF" in val:
                label_map["cpf"] = row_idx
            elif "TEL" in val:
                label_map["tel"] = row_idx
            elif "EMAIL" in val:
                label_map["email"] = row_idx
            elif "APELIDO" in val:
                label_map["apelido"] = row_idx
            elif "TIPO" in val or "T.FOR" in val:
                label_map["tipo"] = row_idx
            elif "ENDERE" in val:
                label_map["endereco"] = row_idx
            elif "BAIRRO" in val:
                label_map["bairro"] = row_idx
            elif "CEP" in val:
                label_map["cep"] = row_idx
            elif "CIDADE" in val:
                label_map["cidade"] = row_idx
            elif val == "UC:":
                label_map["uc"] = row_idx
            elif "TARIFA" in val:
                label_map["tarifa"] = row_idx
            elif "SALDO" in val:
                label_map["saldo"] = row_idx

    for col_idx in range(2, ws.max_column + 1):
        uc_val   = _safe_str(ws.cell(row=label_map.get("uc",   19), column=col_idx).value)
        nome_val = _safe_str(ws.cell(row=label_map.get("nome",  2), column=col_idx).value)

        if not uc_val   or uc_val   == "None" or "UC"   in uc_val.upper()   or uc_val.startswith("-"):
            continue
        if not nome_val or nome_val == "None" or "NOME" in nome_val.upper():
            continue

        uc   = uc_val.replace(" ", "").replace(".", "")
        nome = nome_val.upper()

        def _get(field, default=""):
            r = label_map.get(field)
            if not r:
                return default
            v = ws.cell(row=r, column=col_idx).value
            return _safe_str(v) if v is not None else default

        tarifa    = _safe_float(ws.cell(row=label_map["tarifa"], column=col_idx).value if label_map.get("tarifa") else None, 1.125214)
        ja_existe = tb_get_cliente_por_uc(uc) is not None

        cli_salvo = tb_save_cliente({
            "cod_uc":          uc,
            "desc_nome":       nome,
            "desc_cpf":        _get("cpf"),
            "desc_telefone":   _get("tel"),
            "desc_email":      _get("email"),
            "desc_apelido":    _get("apelido"),
            "tp_fornecimento": _tipo_fornecimento(_get("tipo")),
            "pct_desconto":    0.20,
        })
        id_cliente = cli_salvo.get("id_cliente") if cli_salvo else None
        if id_cliente:
            end = {
                "desc_logradouro": _get("endereco").upper(),
                "desc_setor":      _get("bairro").upper(),
                "cod_cep":         _get("cep"),
                "desc_cidade":     _get("cidade"),
            }
            if any(v for v in end.values()):
                tb_save_endereco(id_cliente, end)

        resultados.append({"uc": uc, "nome": nome,
                           "status": "Atualizado" if ja_existe else "Novo", "ok": True})
    return resultados


@bp.route("/importar", methods=["GET", "POST"])
def importar():
    if request.method == "POST":
        if "xlsx" not in request.files or request.files["xlsx"].filename == "":
            flash("Selecione um arquivo Excel!", "danger")
            return redirect(url_for(".importar"))

        xlsx = request.files["xlsx"]
        xlsx_path = os.path.join(UPLOAD_FOLDER, xlsx.filename)
        xlsx.save(xlsx_path)
        formato = request.form.get("formato", "padrao")

        try:
            if formato == "controle":
                resultados = _importar_formato_controle(xlsx_path)
            else:
                resultados = _importar_formato_padrao(xlsx_path)
            return render_template("importar_resultado.html", resultados=resultados)
        except Exception as e:
            traceback.print_exc()
            flash(f"Erro ao importar: {str(e)}", "danger")
            return redirect(url_for(".importar"))

    return render_template("importar.html")


@bp.route("/importar/modelo")
def importar_modelo():
    modelo = os.path.join(_PROJECT_DIR, "CONTALEV_Modelo_Importacao.xlsx")
    if os.path.exists(modelo):
        return send_file(modelo, as_attachment=True)
    flash("Modelo nao encontrado!", "danger")
    return redirect(url_for(".importar"))
