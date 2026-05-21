"""
Gera template Excel para cadastro em lote de USINAS SOLEV.

Estrutura do arquivo:
  - Aba 1: "Usinas" — uma linha por usina com TODOS os campos
  - Aba 2: "Instruções" — guia de preenchimento

Uso:
    python scripts/gerar_template_cadastro_usinas.py

Saída:
    cadastro_usinas_template.xlsx (na raiz do projeto)
"""
import sys, os
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.comments import Comment


# ─── Estrutura: colunas agrupadas por seção ──────────────────────────────
# Cada item: (chave_interna, titulo_visivel, exemplo, descricao, largura)
SECOES = [
    ("🔵 IDENTIFICAÇÃO", "1F4E79", [
        ("desc_nome",        "Nome da Usina*",      "Usina Danilo Evangelista",  "Nome interno (obrigatório)", 28),
        ("cod_uc_geradora",  "UC Geradora*",        "4.355.432.012-70",          "Código da UC no formato xxxx.xxx.xxx.xxx-xx", 22),
        ("desc_classe",      "Classe",              "B1",                         "Classe tarifária (ex: B1, B3)", 8),
        ("qtd_dia_leitura",  "Dia Leitura (1-31)",  "8",                          "Dia HABITUAL da leitura mensal", 14),
        ("dt_proxima_leitura", "Próxima Leitura",   "08/06/2026",                 "Data prevista da próxima leitura (dd/mm/aaaa)", 16),
    ]),
    ("📍 LOCALIZAÇÃO", "548235", [
        ("desc_logradouro",  "Logradouro",          "Rua Principal",              "Rua, Avenida, Estrada", 24),
        ("desc_numero",      "Número",              "S/N",                        "Número ou S/N", 8),
        ("desc_complemento", "Complemento",         "Lote 5, Quadra A",           "Lote, Sala, Apto", 18),
        ("desc_setor",       "Bairro/Setor",        "Zona Rural",                 "Bairro ou setor", 16),
        ("desc_cidade",      "Cidade",              "Trindade",                   "Cidade", 14),
        ("desc_estado",      "UF",                  "GO",                         "Estado (2 letras)", 6),
        ("cod_cep",          "CEP",                 "75.389-386",                 "CEP (com ou sem máscara)", 12),
    ]),
    ("⚡ EQUIPAMENTOS", "C65911", [
        ("qtd_potencia_kwp",         "Potência (kWp)",       "75",          "Potência nominal em kWp", 14),
        ("qtd_modulos",              "Qtd Módulos",          "180",         "Quantidade de painéis", 12),
        ("desc_modulos_tipo",        "Tipo de Módulo",       "Canadian 550W", "Marca e modelo dos painéis", 18),
        ("desc_inversor",            "Inversor",             "Growatt 75kW", "Marca/modelo do inversor", 18),
        ("desc_estrutura",           "Estrutura",            "Solo metálico", "Tipo de fixação", 16),
        ("qtd_geracao_media_mensal", "Geração Média (kWh)",  "14000",       "Geração média mensal estimada", 16),
        ("qtd_geracao_prevista_diaria", "Geração Dia (kWh)", "470",         "Geração média diária estimada", 14),
        ("dt_comissionamento",       "Comissionamento",      "15/03/2025",  "Data de início operação (dd/mm/aaaa)", 14),
        ("desc_garantia_modulos",    "Garantia Módulos",     "25 anos",     "Tempo de garantia dos painéis", 14),
        ("desc_garantia_inversor",   "Garantia Inversor",    "10 anos",     "Tempo de garantia do inversor", 14),
    ]),
    ("👤 TITULAR DA UC", "7030A0", [
        ("titular_nome",          "Titular - Nome",          "JOÃO TITULAR DA SILVA",   "Nome completo do titular da UC", 26),
        ("titular_cpf_cnpj",      "Titular - CPF/CNPJ",      "123.456.789-00",          "CPF ou CNPJ", 16),
        ("titular_telefone",      "Titular - Telefone",      "(62) 99999-9999",         "Celular com DDD", 16),
        ("titular_email",         "Titular - E-mail",        "titular@email.com",       "E-mail de contato", 24),
        ("titular_dt_nascimento", "Titular - Nascimento",    "01/01/1980",              "Data nascimento (dd/mm/aaaa)", 14),
    ]),
    ("🏠 DONO DA USINA", "BF8F00", [
        ("dono_nome",          "Dono - Nome",          "MARIA DONA DOS SANTOS",    "Nome do dono (pode ser igual ao titular)", 26),
        ("dono_cpf_cnpj",      "Dono - CPF/CNPJ",      "987.654.321-00",           "CPF ou CNPJ", 16),
        ("dono_telefone",      "Dono - Telefone",      "(62) 88888-8888",          "Celular com DDD", 16),
        ("dono_email",         "Dono - E-mail",        "dono@email.com",           "E-mail", 24),
        ("dono_dt_nascimento", "Dono - Nascimento",    "10/05/1975",               "Data nascimento", 14),
    ]),
    ("💰 RECEBEDOR PIX", "C00000", [
        ("inv_nome",            "PIX - Nome",          "JOÃO RECEBEDOR",       "Nome titular da conta", 24),
        ("inv_cpf_cnpj",        "PIX - CPF/CNPJ",      "111.222.333-44",       "CPF/CNPJ do titular PIX", 16),
        ("inv_telefone",        "PIX - Telefone",      "(62) 77777-7777",      "Contato", 16),
        ("inv_email",           "PIX - E-mail",        "pix@email.com",        "E-mail do recebedor", 20),
        ("inv_banco",           "Banco",               "Banco do Brasil",      "Nome do banco", 18),
        ("inv_agencia",         "Agência",             "1234-5",               "Agência", 10),
        ("inv_conta",           "Conta",               "12345-6",              "Número da conta", 12),
        ("inv_pix",             "Chave PIX",           "f6189239-d8ae-4edb-9d62-99299de54fc3", "Chave PIX (CPF/email/celular/aleatória)", 38),
        ("inv_pct_desagio",     "Deságio %",           "15",                    "% retido pela SOLEV (ex: 15 = 15%)", 10),
        ("inv_dia_pagamento",   "Dia Pagamento",       "15",                    "Dia do mês de pagamento ao investidor", 12),
        ("inv_valor_minimo",    "Valor Mínimo",        "100",                   "Valor mínimo descontado por fatura (R$)", 12),
    ]),
    ("📝 OBSERVAÇÕES", "404040", [
        ("desc_observacoes",    "Observações",         "Usina instalada em zona rural, acesso por estrada de terra.",  "Notas livres", 50),
    ]),
]

# ─── Cria a planilha ──────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Usinas"

# Estilos
borda_fina = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Cabeçalho de seção (linha 1)
col = 1
for secao_titulo, cor, campos in SECOES:
    largura_secao = len(campos)
    # Mescla células da seção
    ws.merge_cells(
        start_row=1, start_column=col,
        end_row=1, end_column=col + largura_secao - 1
    )
    cell = ws.cell(row=1, column=col, value=secao_titulo)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borda_fina
    col += largura_secao

# Cabeçalho de coluna (linha 2)
col = 1
for secao_titulo, cor, campos in SECOES:
    for chave, titulo, exemplo, descricao, largura in campos:
        cell = ws.cell(row=2, column=col, value=titulo)
        cell.font = Font(bold=True, color="000000", size=10)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_fina
        cell.comment = Comment(descricao, "SOLEV")
        # Largura da coluna
        ws.column_dimensions[get_column_letter(col)].width = largura
        col += 1

# Linha 3: linha de EXEMPLO (preenchida com placeholder)
col = 1
for secao_titulo, cor, campos in SECOES:
    for chave, titulo, exemplo, descricao, largura in campos:
        cell = ws.cell(row=3, column=col, value=exemplo)
        cell.font = Font(italic=True, color="888888", size=10)
        cell.fill = PatternFill("solid", fgColor="FFFFE0")  # amarelo claro
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = borda_fina
        col += 1

# Adiciona algumas linhas vazias formatadas (para o usuário preencher)
for linha in range(4, 14):  # 10 linhas em branco
    col = 1
    for secao_titulo, cor, campos in SECOES:
        for chave, titulo, exemplo, descricao, largura in campos:
            cell = ws.cell(row=linha, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = borda_fina
            col += 1

# Congela primeira linha + primeira coluna (nome da usina) ao rolar
ws.freeze_panes = "B3"

# Altura das linhas
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 32
ws.row_dimensions[3].height = 32

# ─── Aba 2: Instruções ────────────────────────────────────────────────────
ws2 = wb.create_sheet("📋 Instruções")
ws2.column_dimensions['A'].width = 90

instrucoes = [
    ("Como usar este template", True, 14, "1F4E79"),
    ("", False, 10, None),
    ("1. Cada LINHA = uma usina.", False, 11, None),
    ("2. Preencha os campos marcados com * (obrigatórios).", False, 11, None),
    ("3. A LINHA 3 contém um exemplo — pode apagar ou usar como referência.", False, 11, None),
    ("4. Outras linhas (4 em diante): preencha com suas usinas reais.", False, 11, None),
    ("5. Passe o mouse sobre o título de cada coluna pra ver a descrição.", False, 11, None),
    ("", False, 10, None),
    ("Formatos esperados:", True, 12, "548235"),
    ("• Datas: dd/mm/aaaa (ex: 08/06/2026)", False, 11, None),
    ("• Valores numéricos: use vírgula como decimal (ex: 14.143,50 ou 14143,50)", False, 11, None),
    ("• UC: 15 dígitos no formato xxxx.xxx.xxx.xxx-xx ou apenas números", False, 11, None),
    ("• CEP: xx.xxx-xxx ou apenas dígitos", False, 11, None),
    ("• Telefone: (xx) xxxxx-xxxx", False, 11, None),
    ("• CPF: xxx.xxx.xxx-xx | CNPJ: xx.xxx.xxx/xxxx-xx", False, 11, None),
    ("", False, 10, None),
    ("Dicas práticas:", True, 12, "C65911"),
    ("• Se o TITULAR e o DONO forem a mesma pessoa, repita os dados em ambas as seções.", False, 11, None),
    ("• Se ainda não souber a 'Próxima Leitura', deixe em branco — pode ajustar depois.", False, 11, None),
    ("• O 'Dia Leitura' (1-31) é o dia HABITUAL — usado pra calcular o deadline de rateio (D-7).", False, 11, None),
    ("• 'Deságio %' do PIX = porcentagem retida pela SOLEV antes de pagar o investidor.", False, 11, None),
    ("• Campos em branco serão ignorados na importação.", False, 11, None),
    ("", False, 10, None),
    ("Após preencher:", True, 12, "7030A0"),
    ("• Salve o arquivo (Ctrl+S).", False, 11, None),
    ("• Envie pro Claude (ou rode o script de importação quando estiver pronto).", False, 11, None),
    ("• Verifique no painel de Usinas (/usinas) que todas foram criadas corretamente.", False, 11, None),
    ("", False, 10, None),
    ("Em caso de dúvida sobre um campo específico, deixe em branco e ajuste no painel depois.", True, 11, "C00000"),
]

for i, (texto, bold, size, cor) in enumerate(instrucoes, start=1):
    c = ws2.cell(row=i, column=1, value=texto)
    c.font = Font(bold=bold, size=size, color=cor or "000000")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws2.row_dimensions[i].height = max(18, size + 8)

# Salva
out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "cadastro_usinas_template.xlsx")
wb.save(out_path)
print(f"\n✓ Template gerado: {out_path}")
print(f"  Tamanho: {os.path.getsize(out_path):,} bytes")
print(f"  Colunas: {sum(len(c) for _, _, c in SECOES)} campos em {len(SECOES)} seções")
print(f"\nAbra com Excel ou LibreOffice e comece a preencher!")
