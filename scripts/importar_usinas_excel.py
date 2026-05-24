"""
Importa usinas do arquivo cadastro_usinas_template.xlsx para o Supabase.

Faz dedup automaticamente:
  - Titulares por CPF/CNPJ (vários nomes no Excel podem ser a mesma pessoa)
  - Donos por CPF/CNPJ
  - Investidores (PIX) por CPF/CNPJ
  - Usinas por cod_uc_geradora (evita re-cadastrar a Danilo Evangelista que já existe)

Uso:
    python scripts/importar_usinas_excel.py            # dry-run (mostra o que faria)
    python scripts/importar_usinas_excel.py --aplicar  # executa de fato
"""
import sys, os, re
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv; load_dotenv()
from openpyxl import load_workbook
from db import _db, tb_save_titular, tb_save_dono, tb_save_investidor, tb_save_endereco_usina
from datetime import datetime

DRY_RUN = "--aplicar" not in sys.argv

XLSX = "cadastro_usinas_template.xlsx"

# Limpa CPF/CNPJ (só dígitos) — usado pra dedup
def _digitos(s):
    if s is None: return ""
    return re.sub(r"\D", "", str(s))

# Normaliza UC (só dígitos) — usado pra dedup
def _norm_uc(s):
    return _digitos(s).lstrip("0")

# Converte qualquer formato de data pra ISO 'YYYY-MM-DD'
def _data_iso(v):
    if v is None or v == "": return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try: return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError: continue
    return None

# Extrai dia (1-31) de uma data ou número
def _extrai_dia(v):
    if v is None or v == "": return None
    if isinstance(v, datetime): return v.day
    try:
        n = int(str(v).strip())
        if 1 <= n <= 31: return n
    except (ValueError, TypeError): pass
    return None


def main():
    print(f"{'DRY-RUN (sem aplicar)' if DRY_RUN else 'APLICANDO'}")
    print("=" * 70)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Usinas"]

    db = _db()

    # Carrega o que já existe pra dedup
    print("\n1) Carregando estado atual...")
    titulares_existentes = {_digitos(t.get("desc_cpf_cnpj")): t
                            for t in (db.select("tb_titulares") or [])
                            if t.get("desc_cpf_cnpj")}
    donos_existentes = {_digitos(d.get("desc_cpf_cnpj")): d
                        for d in (db.select("tb_donos") or [])
                        if d.get("desc_cpf_cnpj")}
    investidores_existentes = {_digitos(i.get("desc_cpf_cnpj")): i
                               for i in (db.select("tb_investidores") or [])
                               if i.get("desc_cpf_cnpj")}
    usinas_existentes = {_norm_uc(u.get("cod_uc_geradora")): u
                         for u in (db.select("tb_usinas") or [])
                         if u.get("cod_uc_geradora")}
    print(f"   Titulares: {len(titulares_existentes)}")
    print(f"   Donos:     {len(donos_existentes)}")
    print(f"   Investidores: {len(investidores_existentes)}")
    print(f"   Usinas:    {len(usinas_existentes)}")

    # Lê linhas (4 em diante)
    print("\n2) Lendo Excel...")
    linhas = []
    for row in range(4, ws.max_row + 1):
        nome = ws.cell(row=row, column=1).value
        if not nome: continue
        def c(col): return ws.cell(row=row, column=col).value
        linhas.append({
            "nome":         c(1), "cod_uc":  c(2), "classe":  c(3),
            "dia_leitura":  c(4), "prox_leit": c(5),
            "logradouro":   c(6), "numero":  c(7), "complem": c(8),
            "setor":  c(9), "cidade": c(10), "uf": c(11), "cep": c(12),
            "potencia":     c(13), "qtd_modulos": c(14), "tipo_modulo": c(15),
            "inversor":     c(16), "estrutura":  c(17),
            "ger_media":    c(18), "ger_dia":    c(19),
            "comiss":       c(20), "gar_modulos": c(21), "gar_inv": c(22),
            "tit_nome":     c(23), "tit_cpf":  c(24), "tit_tel":  c(25), "tit_email": c(26), "tit_nasc": c(27),
            "dono_nome":    c(28), "dono_cpf": c(29), "dono_tel": c(30), "dono_email": c(31), "dono_nasc": c(32),
            "pix_nome":     c(33), "pix_cpf":  c(34), "pix_tel":  c(35), "pix_email": c(36),
            "banco":        c(37), "agencia":  c(38), "conta":  c(39), "pix_chave": c(40),
            "desagio":      c(41), "dia_pgto": c(42), "vlr_min": c(43),
            "obs":          c(44),
        })
    print(f"   {len(linhas)} usinas lidas do Excel")

    # Processa cada linha
    print("\n3) Processando usinas...")
    stats = {"titulares_novos": 0, "donos_novos": 0, "investidores_novos": 0,
             "usinas_novas": 0, "usinas_atualizadas": 0, "enderecos": 0, "erros": 0}

    for i, l in enumerate(linhas, 1):
        print(f"\n[{i}/{len(linhas)}] {l['nome']}  UC={l['cod_uc']}")
        try:
            # ---- Titular ----
            id_titular = None
            tit_cpf_d = _digitos(l["tit_cpf"])
            if l["tit_nome"]:
                if tit_cpf_d and tit_cpf_d in titulares_existentes:
                    id_titular = titulares_existentes[tit_cpf_d]["id_titular"]
                    print(f"  Titular (existente): {l['tit_nome'][:40]}  id={id_titular}")
                else:
                    dados = {"desc_nome": l["tit_nome"],
                             "desc_cpf_cnpj": l["tit_cpf"],
                             "desc_telefone": l["tit_tel"],
                             "desc_email": l["tit_email"],
                             "dt_nascimento": _data_iso(l["tit_nasc"])}
                    if not DRY_RUN:
                        t = tb_save_titular(dados)
                        id_titular = t.get("id_titular")
                        if tit_cpf_d: titulares_existentes[tit_cpf_d] = t
                    else:
                        # DRY-RUN: registra placeholder pra dedup intra-Excel
                        if tit_cpf_d: titulares_existentes[tit_cpf_d] = {"id_titular": f"PLACEHOLDER_{tit_cpf_d}", **dados}
                    stats["titulares_novos"] += 1
                    print(f"  Titular (NOVO):     {l['tit_nome'][:40]}")

            # ---- Dono ----
            id_dono = None
            dono_cpf_d = _digitos(l["dono_cpf"])
            if l["dono_nome"]:
                if dono_cpf_d and dono_cpf_d in donos_existentes:
                    id_dono = donos_existentes[dono_cpf_d]["id_dono"]
                    print(f"  Dono (existente):   {l['dono_nome'][:40]}  id={id_dono}")
                else:
                    dados = {"desc_nome": l["dono_nome"],
                             "desc_cpf_cnpj": l["dono_cpf"],
                             "desc_telefone": l["dono_tel"],
                             "desc_email": l["dono_email"],
                             "dt_nascimento": _data_iso(l["dono_nasc"])}
                    if not DRY_RUN:
                        d = tb_save_dono(dados)
                        id_dono = d.get("id_dono")
                        if dono_cpf_d: donos_existentes[dono_cpf_d] = d
                    else:
                        if dono_cpf_d: donos_existentes[dono_cpf_d] = {"id_dono": f"PLACEHOLDER_{dono_cpf_d}", **dados}
                    stats["donos_novos"] += 1
                    print(f"  Dono (NOVO):        {l['dono_nome'][:40]}")

            # ---- Investidor (PIX) ----
            id_investidor = None
            pix_cpf_d = _digitos(l["pix_cpf"])
            if l["pix_nome"] and l["pix_chave"]:
                if pix_cpf_d and pix_cpf_d in investidores_existentes:
                    id_investidor = investidores_existentes[pix_cpf_d]["id_investidor"]
                    print(f"  PIX (existente):    {l['pix_nome'][:40]}  id={id_investidor}")
                else:
                    dados = {"desc_nome": l["pix_nome"],
                             "desc_cpf_cnpj": l["pix_cpf"],
                             "desc_telefone": l["pix_tel"],
                             "desc_email": l["pix_email"],
                             "desc_banco": l["banco"],
                             "desc_agencia": l["agencia"],
                             "desc_conta": l["conta"],
                             "desc_pix": l["pix_chave"],
                             "pct_desagio": float(l["desagio"] or 0) if l["desagio"] else 0,
                             "qtd_dia_pagamento": int(l["dia_pgto"] or 0) if l["dia_pgto"] else None,
                             "vlr_minimo": float(l["vlr_min"] or 0) if l["vlr_min"] else 0}
                    if not DRY_RUN:
                        i_obj = tb_save_investidor(dados)
                        id_investidor = i_obj.get("id_investidor")
                        if pix_cpf_d: investidores_existentes[pix_cpf_d] = i_obj
                    else:
                        if pix_cpf_d: investidores_existentes[pix_cpf_d] = {"id_investidor": f"PLACEHOLDER_{pix_cpf_d}", **dados}
                    stats["investidores_novos"] += 1
                    print(f"  PIX (NOVO):         {l['pix_nome'][:40]}")

            # ---- Usina ----
            # cod_uc_geradora exige exatos 15 dígitos sem formatação (check constraint)
            uc_digitos = _digitos(l["cod_uc"])
            uc_norm = uc_digitos.lstrip("0")
            usina_existente = usinas_existentes.get(uc_norm)
            dados_usina = {
                "desc_nome":                  l["nome"],
                "cod_uc_geradora":            uc_digitos,
                "desc_classe":                l["classe"],
                "qtd_potencia_kwp":           float(l["potencia"] or 0) if l["potencia"] else 0,
                "qtd_modulos":                int(l["qtd_modulos"] or 0) if l["qtd_modulos"] else 0,
                "desc_modulos_tipo":          l["tipo_modulo"],
                "desc_inversor":              l["inversor"],
                "desc_estrutura":             l["estrutura"],
                "qtd_geracao_media_mensal":   float(l["ger_media"] or 0) if l["ger_media"] else 0,
                "qtd_geracao_prevista_diaria":float(l["ger_dia"] or 0) if l["ger_dia"] else 0,
                "dt_comissionamento":         _data_iso(l["comiss"]),
                "desc_garantia_modulos":      l["gar_modulos"] or "25 anos",
                "desc_garantia_inversor":     l["gar_inv"] or "10 anos",
                "qtd_dia_leitura":            _extrai_dia(l["dia_leitura"]),
                "dt_proxima_leitura":         _data_iso(l["prox_leit"]),
                "desc_observacoes":           l["obs"] or "",
            }
            if id_titular: dados_usina["id_titular"] = id_titular
            if id_dono: dados_usina["id_dono"] = id_dono
            if id_investidor: dados_usina["id_investidor"] = id_investidor

            id_usina = None
            if usina_existente:
                id_usina = usina_existente["id_usina"]
                dados_usina["id_usina"] = id_usina
                if not DRY_RUN:
                    db.upsert_returning("tb_usinas", dados_usina, on_conflict="id_usina")
                stats["usinas_atualizadas"] += 1
                print(f"  Usina (atualizada): id={id_usina}")
            else:
                if not DRY_RUN:
                    u = db.upsert_returning("tb_usinas", dados_usina, on_conflict=None)
                    id_usina = u.get("id_usina")
                    usinas_existentes[uc_norm] = u
                stats["usinas_novas"] += 1
                print(f"  Usina (NOVA)" + (f" id={id_usina}" if id_usina else ""))

            # ---- Endereço da usina ----
            if id_usina and (l["logradouro"] or l["cidade"]):
                end = {
                    "desc_logradouro":  l["logradouro"],
                    "desc_numero":      l["numero"],
                    "desc_complemento": l["complem"],
                    "desc_setor":       l["setor"],
                    "desc_cidade":      l["cidade"],
                    "desc_estado":      l["uf"],
                    "cod_cep":          l["cep"],
                }
                if not DRY_RUN:
                    tb_save_endereco_usina(id_usina, end)
                stats["enderecos"] += 1
                print(f"  Endereço:           {l['cidade']}/{l['uf']}")

        except Exception as e:
            stats["erros"] += 1
            import traceback
            print(f"  ERRO: {e}")
            traceback.print_exc()

    # Resumo
    print("\n" + "=" * 70)
    print(f"RESUMO ({'DRY-RUN' if DRY_RUN else 'APLICADO'}):")
    print(f"  Titulares novos:      {stats['titulares_novos']}")
    print(f"  Donos novos:          {stats['donos_novos']}")
    print(f"  Investidores novos:   {stats['investidores_novos']}")
    print(f"  Usinas novas:         {stats['usinas_novas']}")
    print(f"  Usinas atualizadas:   {stats['usinas_atualizadas']}")
    print(f"  Endereços salvos:     {stats['enderecos']}")
    print(f"  Erros:                {stats['erros']}")
    if DRY_RUN:
        print("\nPra aplicar de fato, rode:")
        print("  python scripts/importar_usinas_excel.py --aplicar")


if __name__ == "__main__":
    main()
