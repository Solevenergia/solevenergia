"""
Gera planilha SOLEV_Preenchimento.xlsx com:
- Aba "Clientes": id, nome, UC, consumo médio, saldo atual, próxima leitura, dia leitura, status
- Aba "Usinas":   id, nome, UC, dia de leitura, próxima leitura, geração média, PIX recebimento
Campos faltantes ficam em amarelo para o usuário preencher.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import (tb_carregar_clientes, tb_carregar_usinas, _db)

# ─── Estilos ─────────────────────────────────────────────────────────
FONT_TITULO  = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_HEADER  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_SUB     = Font(name="Arial", size=9, italic=True, color="666666")
FONT_BODY    = Font(name="Arial", size=10)
FONT_BODY_BD = Font(name="Arial", size=10, bold=True)

FILL_TITULO = PatternFill("solid", fgColor="1A2A4A")   # azul SOLEV
FILL_HEADER = PatternFill("solid", fgColor="2A4068")
FILL_ALT    = PatternFill("solid", fgColor="F5F7FA")
FILL_FALTA  = PatternFill("solid", fgColor="FFF3CD")   # amarelo: faltante
FILL_OK     = PatternFill("solid", fgColor="D4EDDA")   # verde claro: preenchido
FILL_LOCK   = PatternFill("solid", fgColor="E9ECEF")   # cinza: não editar (id)

BORDER = Border(
    left=Side(style="thin",  color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin",   color="CCCCCC"),
    bottom=Side(style="thin",color="CCCCCC"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


def _falta(v):
    """Retorna True se valor estiver vazio/zero (precisa preencher)."""
    if v is None or v == "":
        return True
    try:
        return float(v) == 0
    except (TypeError, ValueError):
        return False


def _data_br(iso: str) -> str:
    """ISO YYYY-MM-DD → dd/mm/aaaa."""
    if not iso or len(str(iso)) < 10:
        return ""
    s = str(iso)[:10]
    return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"


def _aplica_header(ws, headers: list, row: int = 3):
    """Escreve cabeçalho com subtítulos opcionais (tuple = (titulo, sub))."""
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        if isinstance(h, tuple):
            cell.value = h[0]
            # subtítulo na linha abaixo
            sub = ws.cell(row=row + 1, column=col)
            sub.value = h[1]
            sub.font = FONT_SUB
            sub.fill = FILL_HEADER
            sub.alignment = ALIGN_CENTER
            sub.border = BORDER
        else:
            cell.value = h
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER


def _aba_clientes(wb: Workbook):
    ws = wb.create_sheet("Clientes")

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"] = "SOLEV — Dados dos Clientes (preencha os campos amarelos)"
    ws["A1"].font = FONT_TITULO
    ws["A1"].fill = FILL_TITULO
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    # Cabeçalho (2 linhas)
    headers = [
        ("ID",                "(não alterar)"),
        ("Nome",              ""),
        ("UC",                "15 dígitos"),
        ("Consumo Médio",     "kWh/mês"),
        ("Saldo Atual",       "kWh"),
        ("Próxima Leitura",   "dd/mm/aaaa"),
        ("Dia Habitual",      "(opcional: dia do mês)"),
        ("Observações",       ""),
    ]
    _aplica_header(ws, headers, row=3)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18

    # Busca clientes + saldo atual (vinculações ativas)
    clientes = tb_carregar_clientes()
    clientes = [c for c in clientes if str(c.get("STATUS") or "").upper() != "INATIVO"]
    clientes.sort(key=lambda c: (c.get("desc_nome") or "").upper())

    vinculos = _db().select("tb_cliente_usina",
                            raw_params={"dt_fim": "is.null"},
                            columns="id_cliente,qtd_saldo_kwh")
    saldo_map = {}
    for v in vinculos:
        id_c = v.get("id_cliente")
        s = float(v.get("qtd_saldo_kwh") or 0)
        if s > 0:
            saldo_map[id_c] = saldo_map.get(id_c, 0) + s

    # Linhas
    row = 5
    for i, c in enumerate(clientes):
        id_c   = c.get("id_cliente")
        nome   = c.get("desc_nome") or c.get("desc_apelido") or ""
        uc     = c.get("cod_uc") or ""
        consumo = c.get("qtd_consumo_medio_kwh")
        saldo  = saldo_map.get(id_c, c.get("qtd_saldo_inicial_kwh") or 0)
        prox   = _data_br(c.get("proxima_leitura"))
        dia    = int(str(c.get("proxima_leitura") or "")[8:10]) if c.get("proxima_leitura") and len(str(c.get("proxima_leitura"))) >= 10 else ""

        valores = [id_c, nome, uc, consumo or "", saldo or "", prox, dia, ""]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.border = BORDER
            # alinhamento
            if col_idx in (1, 7):
                cell.alignment = ALIGN_CENTER
            elif col_idx in (4, 5):
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT
            # cor por campo
            if col_idx == 1:
                cell.fill = FILL_LOCK
                cell.font = FONT_BODY_BD
            elif col_idx in (4, 5, 6) and _falta(val):
                cell.fill = FILL_FALTA  # amarelo: faltante
            elif col_idx in (4, 5, 6):
                cell.fill = FILL_OK     # verde: preenchido
            elif i % 2 == 0:
                cell.fill = FILL_ALT
        row += 1

    # Larguras
    larguras = [8, 35, 22, 16, 14, 16, 14, 28]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A5"

    # Resumo no topo direito
    ws["J1"] = f"Total: {len(clientes)}"
    ws["J1"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    ws["J1"].fill = FILL_TITULO
    ws["J1"].alignment = ALIGN_CENTER
    falta_consumo = sum(1 for c in clientes if _falta(c.get("qtd_consumo_medio_kwh")))
    falta_leitura = sum(1 for c in clientes if not c.get("proxima_leitura"))
    ws["J2"] = f"Sem consumo: {falta_consumo} | Sem leitura: {falta_leitura}"
    ws["J2"].font = FONT_SUB

    return ws


def _aba_usinas(wb: Workbook):
    ws = wb.create_sheet("Usinas")

    ws.merge_cells("A1:H1")
    ws["A1"] = "SOLEV — Dados das Usinas (preencha os campos amarelos)"
    ws["A1"].font = FONT_TITULO
    ws["A1"].fill = FILL_TITULO
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    headers = [
        ("ID",                  "(não alterar)"),
        ("Nome da Usina",       ""),
        ("UC Geradora",         "15 dígitos"),
        ("Potência",            "kWp"),
        ("Dia Leitura",         "(1-31)"),
        ("Próxima Leitura",     "dd/mm/aaaa"),
        ("Geração Média",       "kWh/mês"),
        ("PIX Recebimento",     "chave SOLEV"),
    ]
    _aplica_header(ws, headers, row=3)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18

    usinas = tb_carregar_usinas()
    usinas.sort(key=lambda u: (u.get("desc_nome") or "").upper())

    row = 5
    for i, u in enumerate(usinas):
        id_u    = u.get("id_usina")
        nome    = u.get("desc_nome") or ""
        uc      = u.get("cod_uc_geradora") or ""
        pot     = u.get("qtd_potencia_kwp") or ""
        dia     = u.get("qtd_dia_leitura") or ""
        prox    = _data_br(u.get("dt_proxima_leitura"))
        ger     = u.get("qtd_geracao_media_mensal") or ""
        pix     = u.get("desc_pix_recebimento") or ""

        valores = [id_u, nome, uc, pot, dia, prox, ger, pix]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_BODY
            cell.border = BORDER
            if col_idx in (1, 5, 6):
                cell.alignment = ALIGN_CENTER
            elif col_idx in (4, 7):
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '#,##0.00' if col_idx == 4 else '#,##0'
            else:
                cell.alignment = ALIGN_LEFT
            if col_idx == 1:
                cell.fill = FILL_LOCK
                cell.font = FONT_BODY_BD
            elif col_idx in (5, 6, 7, 8) and _falta(val):
                cell.fill = FILL_FALTA
            elif col_idx in (5, 6, 7, 8):
                cell.fill = FILL_OK
            elif i % 2 == 0:
                cell.fill = FILL_ALT
        row += 1

    larguras = [8, 32, 22, 12, 12, 16, 14, 36]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    ws["J1"] = f"Total: {len(usinas)}"
    ws["J1"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    ws["J1"].fill = FILL_TITULO
    ws["J1"].alignment = ALIGN_CENTER
    falta_dia = sum(1 for u in usinas if not u.get("qtd_dia_leitura"))
    falta_ger = sum(1 for u in usinas if _falta(u.get("qtd_geracao_media_mensal")))
    ws["J2"] = f"Sem dia leitura: {falta_dia} | Sem geração: {falta_ger}"
    ws["J2"].font = FONT_SUB

    return ws


def _aba_instrucoes(wb: Workbook):
    ws = wb.create_sheet("Instruções", 0)  # primeira aba

    ws.merge_cells("A1:E1")
    ws["A1"] = "SOLEV — Planilha de Preenchimento de Dados"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = FILL_TITULO
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 36

    blocos = [
        ("Como usar", [
            "1. Abra as abas 'Clientes' e 'Usinas' nos cantos inferiores.",
            "2. Os campos AMARELOS ainda não foram preenchidos no sistema — preencha-os.",
            "3. Os campos VERDES já estão preenchidos no sistema — você pode revisar se quiser.",
            "4. NÃO altere a coluna ID (cinza) — ela é usada para identificar o registro.",
            "5. Salve a planilha e envie de volta. O sistema importa os dados novos.",
        ]),
        ("Campos da aba 'Clientes'", [
            "• Consumo Médio (kWh/mês): a média mensal de consumo dos últimos meses (vista na fatura Equatorial).",
            "• Saldo Atual (kWh): saldo de energia que o cliente tem acumulado.",
            "• Próxima Leitura (dd/mm/aaaa): data prevista da próxima leitura da Equatorial.",
            "• Dia Habitual: dia do mês em que a Equatorial faz a leitura (calculado da data acima).",
        ]),
        ("Campos da aba 'Usinas'", [
            "• Dia Leitura (1-31): dia HABITUAL da leitura da usina pela Equatorial.",
            "• Próxima Leitura: data exata da próxima leitura (atualizada quando sobe fatura).",
            "• Geração Média (kWh/mês): produção mensal estimada da usina.",
            "• PIX Recebimento: chave PIX da SOLEV para esta usina (clientes pagam aqui).",
        ]),
        ("Regra de distribuição", [
            "Para o sistema distribuir clientes nas usinas automaticamente:",
            "• O cliente deve ser lido em até 7 dias APÓS a leitura da usina.",
            "• Saldo do cliente reduz a demanda dele na geração da usina.",
            "• O sistema prefere alocar 1 cliente em 1 usina, mas pode dividir em 2 se necessário.",
        ]),
    ]

    row = 3
    for titulo, itens in blocos:
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1, value=titulo)
        c.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 22
        row += 1
        for item in itens:
            ws.merge_cells(f"A{row}:E{row}")
            c = ws.cell(row=row, column=1, value=item)
            c.font = FONT_BODY
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            ws.row_dimensions[row].height = 20
            row += 1
        row += 1  # espaço entre blocos

    # Legenda de cores
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="Legenda de cores")
    c.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    c.fill = FILL_HEADER
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[row].height = 22
    row += 1
    legendas = [
        (FILL_FALTA, "AMARELO — campo vazio, precisa preencher"),
        (FILL_OK,    "VERDE — campo já preenchido (revise se quiser)"),
        (FILL_LOCK,  "CINZA — não alterar (ID do registro)"),
    ]
    for fill, txt in legendas:
        c1 = ws.cell(row=row, column=1, value=" ")
        c1.fill = fill
        c1.border = BORDER
        ws.merge_cells(f"B{row}:E{row}")
        c2 = ws.cell(row=row, column=2, value=txt)
        c2.font = FONT_BODY
        c2.alignment = ALIGN_LEFT
        row += 1

    # Larguras
    ws.column_dimensions["A"].width = 4
    for col in "BCDE":
        ws.column_dimensions[col].width = 24


def main():
    output = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          "SOLEV_Preenchimento.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # remove "Sheet" padrão

    _aba_instrucoes(wb)
    _aba_clientes(wb)
    _aba_usinas(wb)

    wb.save(output)
    print(f"[OK] Planilha gerada: {output}")
    print(f"     Tamanho: {os.path.getsize(output) / 1024:.1f} KB")


if __name__ == "__main__":
    main()
